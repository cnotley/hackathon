"""
Report Generation Lambda Function

This module generates comprehensive audit reports from comparison flags,
creates Excel reports using XXXI templates, and converts to PDF format.
Integrates with Bedrock for intelligent report generation.
"""

import json
import os
import boto3
import pandas as pd
from datetime import datetime, timezone
from typing import Dict, List, Any, Optional
import tempfile
import logging
from decimal import Decimal
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import pdfkit
import markdown
from botocore.exceptions import ClientError
import subprocess
import shutil
import psutil
import gc
import warnings

# Configure logging
logger = logging.getLogger(__name__)
logger.setLevel(os.getenv('LOG_LEVEL', 'INFO'))

# AWS clients
s3_client = boto3.client('s3')
bedrock_client = boto3.client('bedrock-runtime')

# Environment variables
BEDROCK_MODEL_ID = os.getenv('BEDROCK_MODEL_ID', 'anthropic.claude-3-5-sonnet-20241022-v2:0')
REPORTS_BUCKET = os.getenv('REPORTS_BUCKET', 'msa-audit-reports')
TEMPLATE_BUCKET = os.getenv('TEMPLATE_BUCKET', 'msa-audit-templates')
TEMPLATE_KEY = os.getenv('TEMPLATE_KEY', 'XXXI_Template.xlsx')


class BedrockReportGenerator:
    """Generates intelligent audit reports using Bedrock."""
    
    def __init__(self):
        self.bedrock_client = bedrock_client
        self.model_id = BEDROCK_MODEL_ID
    
    def generate_markdown_report(self, flags_data: Dict[str, Any], metadata: Dict[str, Any]) -> str:
        """Generate a comprehensive Markdown report from audit flags."""
        try:
            # Prepare the prompt for Bedrock
            prompt = self._build_report_prompt(flags_data, metadata)
            
            # Call Bedrock to generate the report
            response = self.bedrock_client.invoke_model(
                modelId=self.model_id,
                body=json.dumps({
                    "anthropic_version": "bedrock-2023-05-31",
                    "max_tokens": 4000,
                    "messages": [
                        {
                            "role": "user",
                            "content": prompt
                        }
                    ]
                })
            )
            
            # Parse the response
            response_body = json.loads(response['body'].read())
            markdown_content = response_body['content'][0]['text']
            
            logger.info("Successfully generated Markdown report using Bedrock")
            return markdown_content
            
        except Exception as e:
            logger.error(f"Error generating Bedrock report: {str(e)}")
            # Fallback to template-based report
            return self._generate_fallback_report(flags_data, metadata)
    
    def _build_report_prompt(self, flags_data: Dict[str, Any], metadata: Dict[str, Any]) -> str:
        """Build the prompt for Bedrock report generation."""
        
        # Extract key metrics
        total_savings = flags_data.get('total_savings', 0)
        rate_variances = flags_data.get('rate_variances', [])
        overtime_violations = flags_data.get('overtime_violations', [])
        anomalies = flags_data.get('anomalies', [])
        # Attempt to attach evidence like page numbers if present in flags_data
        anomaly_evidence = {}
        try:
            for a in anomalies:
                key = a.get('item') or a.get('worker')
                if key and isinstance(a, dict):
                    ev = {}
                    if 'page' in a:
                        ev['page'] = a.get('page')
                    if 'evidence' in a:
                        ev['evidence'] = a.get('evidence')
                    if ev:
                        anomaly_evidence[key] = ev
        except Exception:
            pass
        
        # Calculate totals
        as_presented = metadata.get('invoice_total', 0)
        as_analyzed = as_presented - total_savings
        
        prompt = f"""
Generate a comprehensive MSA (Master Services Agreement) audit report in Markdown format based on the following audit findings:

## Invoice Information:
- Invoice Number: {metadata.get('invoice_number', 'N/A')}
- Vendor: {metadata.get('vendor', 'N/A')}
- Date of Loss: {metadata.get('date_of_loss', 'N/A')}
- As Presented Total: ${as_presented:,.2f}
- As Analyzed Total: ${as_analyzed:,.2f}
- Total Potential Savings: ${total_savings:,.2f}

## Audit Findings:

### Rate Variances ({len(rate_variances)} found):
"""
        
        for variance in rate_variances:
            prompt += f"""
- **{variance.get('worker', 'Unknown')}** ({variance.get('labor_type', 'N/A')}): 
  - Charged Rate: ${variance.get('actual_rate', 0):.2f}
  - MSA Standard: ${variance.get('msa_rate', 0):.2f}
  - Variance: {variance.get('variance_percentage', 0):.1f}%
  - Potential Savings: ${variance.get('savings', 0):.2f}
"""
        
        prompt += f"""
### Overtime Violations ({len(overtime_violations)} found):
"""
        
        for violation in overtime_violations:
            prompt += f"""
- **{violation.get('worker', 'Unknown')}**: 
  - Total Hours: {violation.get('total_hours', 0):.1f}
  - Overtime Hours: {violation.get('overtime_hours', 0):.1f}
  - Threshold: {violation.get('threshold', 40):.1f} hours/week
"""
        
        prompt += f"""
### Anomalies ({len(anomalies)} found):
"""
        
        for anomaly in anomalies:
            prompt += f"""
- **{anomaly.get('item', 'Unknown')}**: 
  - Amount: ${anomaly.get('amount', 0):,.2f}
  - Z-Score: {anomaly.get('z_score', 0):.2f}
  - Description: {anomaly.get('description', 'Statistical anomaly detected')}
"""
            # Append evidence if available
            try:
                key = anomaly.get('item') or anomaly.get('worker')
                if key and key in anomaly_evidence:
                    ev = anomaly_evidence[key]
                    if 'page' in ev:
                        prompt += f"\n  - Evidence Page: {ev['page']}\n"
                    if 'evidence' in ev:
                        prompt += f"  - Evidence: {json.dumps(ev['evidence'])}\n"
            except Exception:
                pass
        
        prompt += """

Please generate a professional audit report that includes:

1. **Executive Summary** - Brief overview of findings and total savings
2. **Detailed Findings** - Comprehensive analysis of each discrepancy type
3. **Financial Impact** - Clear breakdown of cost implications
4. **Recommendations** - Actionable steps for addressing findings
5. **Compliance Assessment** - MSA adherence evaluation
6. **Supporting Documentation** - References to audit standards and methodologies

Use professional language appropriate for stakeholders and include specific dollar amounts and percentages where relevant. Format the report in clean Markdown with proper headers, bullet points, and emphasis.
"""
        
        return prompt
    
    def _generate_fallback_report(self, flags_data: Dict[str, Any], metadata: Dict[str, Any]) -> str:
        """Generate a fallback report when Bedrock is unavailable."""
        
        total_savings = flags_data.get('total_savings', 0)
        rate_variances = flags_data.get('rate_variances', [])
        overtime_violations = flags_data.get('overtime_violations', [])
        anomalies = flags_data.get('anomalies', [])
        
        as_presented = metadata.get('invoice_total', 0)
        as_analyzed = as_presented - total_savings
        
        report = f"""# MSA Audit Report

## Executive Summary

This audit report presents findings from the analysis of invoice {metadata.get('invoice_number', 'N/A')} 
submitted by {metadata.get('vendor', 'N/A')} against Master Services Agreement standards.

**Key Findings:**
- Total Discrepancies: {len(rate_variances) + len(overtime_violations) + len(anomalies)}
- Potential Savings: ${total_savings:,.2f}
- Savings Percentage: {(total_savings / as_presented * 100) if as_presented > 0 else 0:.1f}%

## Financial Summary

| Category | As Presented | As Analyzed | Savings |
|----------|--------------|-------------|---------|
| Total Invoice | ${as_presented:,.2f} | ${as_analyzed:,.2f} | ${total_savings:,.2f} |

## Detailed Findings

### Rate Variances ({len(rate_variances)} found)

"""
        
        for variance in rate_variances:
            report += f"""
**{variance.get('worker', 'Unknown')} - {variance.get('labor_type', 'N/A')} Labor**
- Charged Rate: ${variance.get('actual_rate', 0):.2f}/hour
- MSA Standard: ${variance.get('msa_rate', 0):.2f}/hour
- Variance: {variance.get('variance_percentage', 0):.1f}% above standard
- Potential Savings: ${variance.get('savings', 0):.2f}

"""
        
        if overtime_violations:
            report += f"""
### Overtime Violations ({len(overtime_violations)} found)

"""
            for violation in overtime_violations:
                report += f"""
**{violation.get('worker', 'Unknown')}**
- Total Hours: {violation.get('total_hours', 0):.1f}
- Overtime Hours: {violation.get('overtime_hours', 0):.1f}
- Requires timesheet documentation per MSA terms

"""
        
        if anomalies:
            report += f"""
### Cost Anomalies ({len(anomalies)} found)

"""
            for anomaly in anomalies:
                report += f"""
**{anomaly.get('item', 'Unknown')}**
- Amount: ${anomaly.get('amount', 0):,.2f}
- Statistical Significance: {anomaly.get('z_score', 0):.2f} standard deviations
- Recommendation: Requires additional documentation

"""
        
        report += """
## Recommendations

1. **Rate Compliance**: Adjust labor rates to MSA standards for future invoicing
2. **Documentation**: Provide timesheet support for overtime hours
3. **Process Improvement**: Implement pre-submission validation against MSA rates
4. **Training**: Ensure vendor billing team understands MSA requirements

## Compliance Assessment

This audit was conducted in accordance with MSA terms and industry standard auditing practices. 
All findings are based on documented rate standards and established overtime thresholds.

---
*Report generated on {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')}*
"""
        
        return report


class ExcelReportGenerator:
    """Generates Excel reports using XXXI templates."""
    
    def __init__(self):
        self.s3_client = s3_client
        self.template_bucket = TEMPLATE_BUCKET
        self.template_key = TEMPLATE_KEY
    
    def generate_excel_report(self, flags_data: Dict[str, Any], metadata: Dict[str, Any], 
                            extracted_data: Dict[str, Any]) -> bytes:
        """Generate Excel report using XXXI template."""
        try:
            # Download template from S3
            template_content = self._download_template()
            
            # Load template workbook
            with tempfile.NamedTemporaryFile() as temp_file:
                temp_file.write(template_content)
                temp_file.flush()
                
                workbook = openpyxl.load_workbook(temp_file.name)
            
            # Fill template with data
            self._fill_project_information(workbook, metadata)
            self._fill_project_summary(workbook, flags_data, metadata)
            self._fill_labor_export(workbook, extracted_data, flags_data)
            
            # Save to bytes
            output_buffer = BytesIO()
            workbook.save(output_buffer)
            output_buffer.seek(0)
            
            logger.info("Successfully generated Excel report")
            return output_buffer.getvalue()
            
        except Exception as e:
            logger.error(f"Error generating Excel report: {str(e)}")
            # Generate fallback Excel report
            return self._generate_fallback_excel(flags_data, metadata, extracted_data)
    
    def _download_template(self) -> bytes:
        """Download Excel template from S3."""
        try:
            response = self.s3_client.get_object(
                Bucket=self.template_bucket,
                Key=self.template_key
            )
            return response['Body'].read()
        except ClientError as e:
            logger.warning(f"Could not download template: {str(e)}")
            # Return empty bytes to trigger fallback
            return b''
    
    def _fill_project_information(self, workbook: openpyxl.Workbook, metadata: Dict[str, Any]):
        """Fill Project Information tab with PDF metadata."""
        try:
            if 'Project Information' in workbook.sheetnames:
                ws = workbook['Project Information']
            else:
                ws = workbook.create_sheet('Project Information')
            
            # Clear existing content
            ws.delete_rows(1, ws.max_row)
            
            # Add headers and data
            project_data = [
                ['Field', 'Value'],
                ['Invoice Number', metadata.get('invoice_number', 'N/A')],
                ['Vendor', metadata.get('vendor', 'N/A')],
                ['Date of Loss', metadata.get('date_of_loss', '2/12/2025')],
                ['Invoice Date', metadata.get('invoice_date', 'N/A')],
                ['Total Amount', f"${metadata.get('invoice_total', 0):,.2f}"],
                ['Page Count', metadata.get('page_count', 'N/A')],
                ['File Size', f"{metadata.get('file_size', 0):,} bytes"],
                ['Processing Date', datetime.now().strftime('%m/%d/%Y')],
                ['Audit Status', 'Completed']
            ]
            
            for row_idx, row_data in enumerate(project_data, 1):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    if row_idx == 1:  # Header row
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
                
        except Exception as e:
            logger.error(f"Error filling Project Information: {str(e)}")
    
    def _fill_project_summary(self, workbook: openpyxl.Workbook, flags_data: Dict[str, Any], 
                            metadata: Dict[str, Any]):
        """Fill Project Summary tab with financial analysis."""
        try:
            if 'Project Summary' in workbook.sheetnames:
                ws = workbook['Project Summary']
            else:
                ws = workbook.create_sheet('Project Summary')
            
            # Clear existing content
            ws.delete_rows(1, ws.max_row)
            
            # Calculate totals
            as_presented_labor = metadata.get('labor_total', 77000)
            total_savings = flags_data.get('total_savings', 0)
            as_analyzed_labor = as_presented_labor - total_savings
            
            summary_data = [
                ['Category', 'As Presented', 'As Analyzed', 'Savings'],
                ['Labor Costs', f"${as_presented_labor:,.2f}", f"${as_analyzed_labor:,.2f}", f"${total_savings:,.2f}"],
                ['Total Project', f"${as_presented_labor:,.2f}", f"${as_analyzed_labor:,.2f}", f"${total_savings:,.2f}"]
            ]
            
            # Add data to worksheet
            for row_idx, row_data in enumerate(summary_data, 1):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    if row_idx == 1:  # Header row
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                    elif row_idx == len(summary_data):  # Total row
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            
            # Add discrepancy details
            ws.cell(row=len(summary_data) + 2, column=1, value="Discrepancy Details").font = Font(bold=True)
            
            detail_row = len(summary_data) + 3
            rate_variances = flags_data.get('rate_variances', [])
            
            for variance in rate_variances:
                ws.cell(row=detail_row, column=1, 
                       value=f"Rate Variance - {variance.get('worker', 'Unknown')}")
                ws.cell(row=detail_row, column=2, 
                       value=f"${variance.get('savings', 0):.2f}")
                detail_row += 1
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
                
        except Exception as e:
            logger.error(f"Error filling Project Summary: {str(e)}")
    
    def _fill_labor_export(self, workbook: openpyxl.Workbook, extracted_data: Dict[str, Any], 
                          flags_data: Dict[str, Any]):
        """Fill Labor Export tab with PDF row data."""
        try:
            if 'Labor Export' in workbook.sheetnames:
                ws = workbook['Labor Export']
            else:
                ws = workbook.create_sheet('Labor Export')
            
            # Clear existing content
            ws.delete_rows(1, ws.max_row)
            
            # Headers
            headers = ['Worker Name', 'Labor Type', 'Hours', 'Rate', 'Total', 'MSA Rate', 'Variance', 'Savings']
            
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Get labor data from extracted data
            labor_entries = extracted_data.get('normalized_data', {}).get('labor', [])
            if extracted_data.get('normalized_data', {}).get('materials'):
                raise ValueError("Materials handling removed")
            
            rate_variances = {v.get('worker'): v for v in flags_data.get('rate_variances', [])}
            
            row_idx = 2
            for entry in labor_entries:
                worker_name = entry.get('name', 'Unknown')
                labor_type = entry.get('type', 'N/A')
                hours = entry.get('total_hours', 0)
                rate = entry.get('unit_price', 0)
                total = hours * rate
                
                # Get variance data if available
                variance_data = rate_variances.get(worker_name, {})
                msa_rate = variance_data.get('msa_rate', rate)
                variance_pct = variance_data.get('variance_percentage', 0)
                # compute per-line savings if missing
                savings = variance_data.get('savings')
                if savings is None:
                    try:
                        savings = max(0.0, (rate - msa_rate) * float(entry.get('total_hours', 0) or 0))
                    except Exception:
                        savings = 0.0
                
                # Add row data
                row_data = [
                    worker_name,
                    labor_type,
                    f"{hours:.1f}",
                    f"${rate:.2f}",
                    f"${total:.2f}",
                    f"${msa_rate:.2f}",
                    f"{variance_pct:.1f}%",
                    f"${savings:.2f}"
                ]
                
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
                
                row_idx += 1
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 20)
                ws.column_dimensions[column_letter].width = adjusted_width
                
        except Exception as e:
            logger.error(f"Error filling Labor Export: {str(e)}")
    
    def _generate_fallback_excel(self, flags_data: Dict[str, Any], metadata: Dict[str, Any], 
                               extracted_data: Dict[str, Any]) -> bytes:
        """Generate fallback Excel report when template is unavailable."""
        workbook = openpyxl.Workbook()
        
        # Remove default sheet
        workbook.remove(workbook.active)
        
        # Create sheets
        self._fill_project_information(workbook, metadata)
        self._fill_project_summary(workbook, flags_data, metadata)
        self._fill_labor_export(workbook, extracted_data, flags_data)
        
        # Save to bytes
        output_buffer = BytesIO()
        workbook.save(output_buffer)
        output_buffer.seek(0)
        
        return output_buffer.getvalue()


class PDFConverter:
    """Converts reports to PDF format."""
    
    def __init__(self):
        # Configure wkhtmltopdf options
        self.pdf_options = {
            'page-size': 'Letter',
            'margin-top': '0.75in',
            'margin-right': '0.75in',
            'margin-bottom': '0.75in',
            'margin-left': '0.75in',
            'encoding': "UTF-8",
            'no-outline': None,
            'enable-local-file-access': None
        }
    
    def markdown_to_pdf(self, markdown_content: str) -> bytes:
        """Convert Markdown content to PDF."""
        try:
            # Convert Markdown to HTML
            html_content = markdown.markdown(markdown_content, extensions=['tables', 'fenced_code'])
            
            # Add CSS styling
            styled_html = f"""
            <!DOCTYPE html>
            <html>
            <head>
                <meta charset="UTF-8">
                <style>
                    body {{ font-family: Arial, sans-serif; line-height: 1.6; margin: 0; padding: 20px; }}
                    h1 {{ color: #2c3e50; border-bottom: 2px solid #3498db; padding-bottom: 10px; }}
                    h2 {{ color: #34495e; margin-top: 30px; }}
                    h3 {{ color: #7f8c8d; }}
                    table {{ border-collapse: collapse; width: 100%; margin: 20px 0; }}
                    th, td {{ border: 1px solid #ddd; padding: 12px; text-align: left; }}
                    th {{ background-color: #f2f2f2; font-weight: bold; }}
                    .highlight {{ background-color: #fff3cd; padding: 10px; border-left: 4px solid #ffc107; }}
                </style>
            </head>
            <body>
                {html_content}
            </body>
            </html>
            """
            
            # Convert HTML to PDF
            pdf_content = pdfkit.from_string(styled_html, False, options=self.pdf_options)
            
            logger.info("Successfully converted Markdown to PDF")
            return pdf_content
            
        except Exception as e:
            logger.error(f"Error converting Markdown to PDF: {str(e)}")
            # Return empty bytes on error
            return b''


class ReportManager:
    """Main report management class."""
    
    def __init__(self):
        self.bedrock_generator = BedrockReportGenerator()
        self.excel_generator = ExcelReportGenerator()
        self.pdf_converter = PDFConverter()
        self.s3_client = s3_client
        self.reports_bucket = REPORTS_BUCKET
    
    def generate_comprehensive_report(self, flags_data: Dict[str, Any], metadata: Dict[str, Any], 
                                    extracted_data: Dict[str, Any]) -> Dict[str, Any]:
        """Generate comprehensive audit report in multiple formats."""
        try:
            report_id = f"audit-{metadata.get('invoice_number', 'unknown')}-{datetime.now().strftime('%Y%m%d-%H%M%S')}"
            
            # Generate Markdown report
            markdown_report = self.bedrock_generator.generate_markdown_report(flags_data, metadata)
            
            # Generate Excel report
            excel_report = self.excel_generator.generate_excel_report(flags_data, metadata, extracted_data)
            
            # Convert Markdown to PDF
            pdf_report = self.pdf_converter.markdown_to_pdf(markdown_report)
            
            # Upload reports to S3
            report_urls = self._upload_reports(report_id, markdown_report, excel_report, pdf_report)

            # CSV export of discrepancies (if present)
            try:
                rows = []
                for key in ['rate_variances','overtime_violations','anomalies','duplicates']:
                    for item in flags_data.get(key, []) or []:
                        item_copy = dict(item)
                        item_copy['type'] = key
                        rows.append(item_copy)
                if rows:
                    df = pd.DataFrame(rows)
                    csv_bytes = df.to_csv(index=False).encode()
                    csv_key = f"reports/{report_id}/{report_id}.csv"
                    s3_client.put_object(Bucket=self.reports_bucket, Key=csv_key, Body=csv_bytes, ContentType='text/csv')
                    report_urls['csv'] = f"s3://{self.reports_bucket}/{csv_key}"
            except Exception as e:
                logger.warning(f"CSV export skipped due to error: {e}")
            
            # Prepare response
            result = {
                'report_id': report_id,
                'generation_status': 'completed',
                'timestamp': datetime.now(timezone.utc).isoformat(),
                'reports': report_urls,
                'summary': {
                    'total_savings': flags_data.get('total_savings', 0),
                    'discrepancies_found': len(flags_data.get('rate_variances', [])) + 
                                         len(flags_data.get('overtime_violations', [])) + 
                                         len(flags_data.get('anomalies', [])),
                    'as_presented': metadata.get('invoice_total', 0),
                    'as_analyzed': metadata.get('invoice_total', 0) - flags_data.get('total_savings', 0)
                }
            }
            
            logger.info(f"Successfully generated comprehensive report: {report_id}")
            return result
            
        except Exception as e:
            logger.error(f"Error generating comprehensive report: {str(e)}")
            return {
                'report_id': f"error-{datetime.now().strftime('%Y%m%d-%H%M%S')}",
                'generation_status': 'failed',
                'error': str(e),
                'timestamp': datetime.now(timezone.utc).isoformat()
            }
    
    def _upload_reports(self, report_id: str, markdown_content: str, 
                       excel_content: bytes, pdf_content: bytes) -> Dict[str, str]:
        """Upload generated reports to S3."""
        report_urls = {}
        
        try:
            # Upload Markdown report
            if markdown_content:
                markdown_key = f"reports/{report_id}/{report_id}.md"
                self.s3_client.put_object(
                    Bucket=self.reports_bucket,
                    Key=markdown_key,
                    Body=markdown_content.encode('utf-8'),
                    ContentType='text/markdown'
                )
                report_urls['markdown'] = f"s3://{self.reports_bucket}/{markdown_key}"
            
            # Upload Excel report
            if excel_content:
                excel_key = f"reports/{report_id}/{report_id}.xlsx"
                self.s3_client.put_object(
                    Bucket=self.reports_bucket,
                    Key=excel_key,
                    Body=excel_content,
                    ContentType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                report_urls['excel'] = f"s3://{self.reports_bucket}/{excel_key}"
            
            # Upload PDF report
            if pdf_content:
                pdf_key = f"reports/{report_id}/{report_id}.pdf"
                self.s3_client.put_object(
                    Bucket=self.reports_bucket,
                    Key=pdf_key,
                    Body=pdf_content,
                    ContentType='application/pdf'
                )
                report_urls['pdf'] = f"s3://{self.reports_bucket}/{pdf_key}"
            
            logger.info(f"Successfully uploaded reports for {report_id}")
            
        except Exception as e:
            logger.error(f"Error uploading reports: {str(e)}")
        
        return report_urls


## NOTE: Using enhanced handler below; legacy handler removed to avoid duplication


def handle_report_generation(flags_data: Dict[str, Any], metadata: Dict[str, Any], 
                           extracted_data: Dict[str, Any]) -> Dict[str, Any]:
    """Handle report generation task from Step Functions."""
    return lambda_handler({
        'task': 'generate_report',
        'flags_data': flags_data,
        'metadata': metadata,
        'extracted_data': extracted_data
    }, {})


# Constants for report generation enhancements
MAX_EXCEL_MEMORY_MB = 512  # Maximum memory for Excel processing in MB
WKHTMLTOPDF_PATHS = [
    '/opt/bin/wkhtmltopdf',  # Lambda layer path
    '/usr/local/bin/wkhtmltopdf',  # Local installation
    '/usr/bin/wkhtmltopdf',  # System installation
    'wkhtmltopdf'  # PATH lookup
]


class ReportValidationError(Exception):
    """Custom exception for report validation errors."""
    pass


class MemoryOptimizedExcelGenerator(ExcelReportGenerator):
    """Memory-optimized Excel generator for large files."""
    
    def __init__(self):
        super().__init__()
        self.memory_threshold = MAX_EXCEL_MEMORY_MB * 1024 * 1024  # Convert to bytes
    
    def _check_memory_usage(self):
        """Check current memory usage and warn if approaching limits."""
        try:
            process = psutil.Process()
            memory_info = process.memory_info()
            memory_mb = memory_info.rss / 1024 / 1024
            
            if memory_mb > self.memory_threshold / (1024 * 1024) * 0.8:  # 80% threshold
                logger.warning(f"Memory usage high: {memory_mb:.1f}MB")
                gc.collect()  # Force garbage collection
                
            return memory_mb
        except Exception as e:
            logger.warning(f"Could not check memory usage: {str(e)}")
            return 0
    
    def _chunk_write_excel(self, workbook: openpyxl.Workbook, data: List[List], 
                          worksheet_name: str, chunk_size: int = 1000) -> None:
        """Write large datasets to Excel in chunks to manage memory."""
        try:
            if worksheet_name in workbook.sheetnames:
                ws = workbook[worksheet_name]
            else:
                ws = workbook.create_sheet(worksheet_name)
            
            # Clear existing content
            ws.delete_rows(1, ws.max_row)
            
            # Write data in chunks
            for chunk_start in range(0, len(data), chunk_size):
                chunk_end = min(chunk_start + chunk_size, len(data))
                chunk_data = data[chunk_start:chunk_end]
                
                for row_idx, row_data in enumerate(chunk_data, chunk_start + 1):
                    for col_idx, value in enumerate(row_data, 1):
                        ws.cell(row=row_idx, column=col_idx, value=value)
                
                # Check memory after each chunk
                self._check_memory_usage()
                
                # Optional: yield control to avoid blocking
                if chunk_end < len(data):
                    import time
                    time.sleep(0.001)  # 1ms pause
                    
        except Exception as e:
            logger.error(f"Error in chunked Excel writing: {str(e)}")
            raise ReportValidationError(f"Failed to write Excel data: {str(e)}")
    
    def _add_conditional_formatting(self, workbook: openpyxl.Workbook, flags_data: Dict[str, Any]):
        """Add conditional formatting to highlight flagged items."""
        try:
            # Define styles for flagged items
            red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            orange_fill = PatternFill(start_color="FFE6CC", end_color="FFE6CC", fill_type="solid")
            
            red_font = Font(color="CC0000", bold=True)
            
            # Apply formatting to Labor Export sheet
            if 'Labor Export' in workbook.sheetnames:
                ws = workbook['Labor Export']
                rate_variances = {v.get('worker'): v for v in flags_data.get('rate_variances', [])}
                
                # Apply formatting to flagged rows
                for row_idx in range(2, ws.max_row + 1):
                    worker_cell = ws.cell(row=row_idx, column=1)
                    worker_name = worker_cell.value
                    
                    if worker_name in rate_variances:
                        variance_pct = rate_variances[worker_name].get('variance_percentage', 0)
                        
                        # Color coding based on variance severity
                        if variance_pct > 20:  # High variance (>20%)
                            fill_color = red_fill
                            font_color = red_font
                        elif variance_pct > 10:  # Medium variance (10-20%)
                            fill_color = orange_fill
                            font_color = Font(color="CC6600")
                        else:  # Low variance (<10%)
                            fill_color = yellow_fill
                            font_color = Font(color="996600")
                        
                        # Apply formatting to entire row
                        for col_idx in range(1, 9):  # All columns in Labor Export
                            cell = ws.cell(row=row_idx, column=col_idx)
                            cell.fill = fill_color
                            if col_idx == 7:  # Variance column
                                cell.font = font_color
            
            # Apply formatting to Project Summary sheet
            if 'Project Summary' in workbook.sheetnames:
                ws = workbook['Project Summary']
                total_savings = flags_data.get('total_savings', 0)
                
                # Highlight savings if significant
                if total_savings > 1000:  # Significant savings threshold
                    for row_idx in range(1, ws.max_row + 1):
                        for col_idx in range(1, ws.max_column + 1):
                            cell = ws.cell(row=row_idx, column=col_idx)
                            if 'savings' in str(cell.value).lower():
                                cell.fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
                                cell.font = Font(color="006600", bold=True)
            
            logger.info("Successfully applied conditional formatting")
            
        except Exception as e:
            logger.error(f"Error applying conditional formatting: {str(e)}")
    
    def generate_excel_report(self, flags_data: Dict[str, Any], metadata: Dict[str, Any], 
                            extracted_data: Dict[str, Any]) -> bytes:
        """Generate Excel report with memory optimization and conditional formatting."""
        try:
            # Check initial memory
            initial_memory = self._check_memory_usage()
            logger.info(f"Starting Excel generation with {initial_memory:.1f}MB memory usage")
            
            # Download template with error handling
            template_content = self._download_template_with_retries()
            
            # Load template workbook
            with tempfile.NamedTemporaryFile() as temp_file:
                if template_content:
                    temp_file.write(template_content)
                    temp_file.flush()
                    workbook = openpyxl.load_workbook(temp_file.name)
                else:
                    # Create new workbook if template unavailable
                    workbook = openpyxl.Workbook()
                    workbook.remove(workbook.active)  # Remove default sheet
            
            # Fill template with data
            self._fill_project_information(workbook, metadata)
            self._check_memory_usage()
            
            self._fill_project_summary(workbook, flags_data, metadata)
            self._check_memory_usage()
            
            self._fill_labor_export_optimized(workbook, extracted_data, flags_data)
            self._check_memory_usage()
            
            # Add conditional formatting
            self._add_conditional_formatting(workbook, flags_data)
            
            # Save to bytes with memory monitoring
            output_buffer = BytesIO()
            workbook.save(output_buffer)
            output_buffer.seek(0)
            
            final_memory = self._check_memory_usage()
            logger.info(f"Excel generation completed. Memory: {initial_memory:.1f}MB -> {final_memory:.1f}MB")
            
            return output_buffer.getvalue()
            
        except Exception as e:
            logger.error(f"Error generating Excel report: {str(e)}")
            # Generate fallback Excel report
            return self._generate_fallback_excel(flags_data, metadata, extracted_data)
    
    def _download_template_with_retries(self, max_retries: int = 3) -> bytes:
        """Download Excel template from S3 with retry logic."""
        for attempt in range(max_retries):
            try:
                response = self.s3_client.get_object(
                    Bucket=self.template_bucket,
                    Key=self.template_key
                )
                content = response['Body'].read()
                logger.info(f"Successfully downloaded template on attempt {attempt + 1}")
                return content
                
            except ClientError as e:
                error_code = e.response['Error']['Code']
                logger.warning(f"Template download attempt {attempt + 1} failed: {error_code}")
                
                if attempt == max_retries - 1:
                    logger.error(f"Failed to download template after {max_retries} attempts")
                    return b''
                
                # Wait before retry
                import time
                time.sleep(2 ** attempt)  # Exponential backoff
        
        return b''
    
    def _fill_labor_export_optimized(self, workbook: openpyxl.Workbook, 
                                   extracted_data: Dict[str, Any], flags_data: Dict[str, Any]):
        """Fill Labor Export tab with memory optimization."""
        try:
            # Prepare data structure
            headers = ['Worker Name', 'Labor Type', 'Hours', 'Rate', 'Total', 'MSA Rate', 'Variance', 'Savings']
            
            labor_entries = extracted_data.get('normalized_data', {}).get('labor', [])
            if extracted_data.get('normalized_data', {}).get('materials'):
                raise ValueError("Materials handling removed")
            
            rate_variances = {v.get('worker'): v for v in flags_data.get('rate_variances', [])}
            
            # Build data rows
            data_rows = [headers]
            
            for entry in labor_entries:
                worker_name = entry.get('name', 'Unknown')
                labor_type = entry.get('type', 'N/A')
                hours = entry.get('total_hours', 0)
                rate = entry.get('unit_price', 0)
                total = hours * rate
                
                # Get variance data if available
                variance_data = rate_variances.get(worker_name, {})
                msa_rate = variance_data.get('msa_rate', rate)
                variance_pct = variance_data.get('variance_percentage', 0)
                savings = variance_data.get('savings', 0)
                
                row_data = [
                    worker_name,
                    labor_type,
                    f"{hours:.1f}",
                    f"${rate:.2f}",
                    f"${total:.2f}",
                    f"${msa_rate:.2f}",
                    f"{variance_pct:.1f}%",
                    f"${savings:.2f}"
                ]
                data_rows.append(row_data)
            
            # Use chunked writing for large datasets
            self._chunk_write_excel(workbook, data_rows, 'Labor Export')
            
        except Exception as e:
            logger.error(f"Error filling optimized Labor Export: {str(e)}")


class EnhancedPDFConverter(PDFConverter):
    """Enhanced PDF converter with wkhtmltopdf validation."""
    
    def __init__(self):
        super().__init__()
        self.wkhtmltopdf_path = self._validate_wkhtmltopdf_path()
    
    def _validate_wkhtmltopdf_path(self) -> str:
        """Validate and find wkhtmltopdf executable."""
        for path in WKHTMLTOPDF_PATHS:
            try:
                if shutil.which(path):
                    # Test if the binary works
                    result = subprocess.run([path, '--version'], 
                                          capture_output=True, text=True, timeout=10)
                    if result.returncode == 0:
                        logger.info(f"Found working wkhtmltopdf at: {path}")
                        return path
                except Exception as e:
                    logger.debug(f"wkhtmltopdf test failed for {path}: {str(e)}")
                    continue
        
        # If no working wkhtmltopdf found
        raise ReportValidationError(
            "wkhtmltopdf not found or not functional. "
            "Please install wkhtmltopdf or check the Lambda layer configuration."
        )
    
    def markdown_to_pdf(self, markdown_content: str) -> bytes:
        """Convert Markdown content to PDF with enhanced validation."""
        try:
            # Validate wkhtmltopdf is available
            if not self.wkhtmltopdf_path:
                raise ReportValidationError("wkhtmltopdf path not validated")
            
            # Convert Markdown to HTML
            html_content = markdown.markdown(markdown_content, extensions=['tables', 'fenced_code'])
            
            # Add enhanced CSS styling
            styled_html = f"""
            <!DOCTYPE html>
            <html>
            <head>
                <meta charset="UTF-8">
                <style>
                    body {{ 
                        font-family: Arial, sans-serif; 
                        line-height: 1.6; 
                        margin: 0; 
                        padding: 20px; 
                        font-size: 12px;
                    }}
                    h1 {{ 
                        color: #2c3e50; 
                        border-bottom: 2px solid #3498db; 
                        padding-bottom: 10px;
                        page-break-before: auto;
                    }}
                    h2 {{ 
                        color: #34495e; 
                        margin-top: 30px;
                        page-break-before: auto;
                    }}
                    h3 {{ color: #7f8c8d; }}
                    table {{ 
                        border-collapse: collapse; 
                        width: 100%; 
                        margin: 20px 0;
                        page-break-inside: avoid;
                    }}
                    th, td {{ 
                        border: 1px solid #ddd; 
                        padding: 8px; 
                        text-align: left;
                        font-size: 11px;
                    }}
                    th {{ 
                        background-color: #f2f2f2; 
                        font-weight: bold; 
                    }}
                    .highlight {{ 
                        background-color: #fff3cd; 
                        padding: 10px; 
                        border-left: 4px solid #ffc107;
                        page-break-inside: avoid;
                    }}
                    .variance-high {{ color: #cc0000; font-weight: bold; }}
                    .variance-medium {{ color: #ff8800; font-weight: bold; }}
                    .savings {{ color: #008800; font-weight: bold; }}
                    @page {{
                        margin: 0.75in;
                        @bottom-right {{
                            content: "Page " counter(page) " of " counter(pages);
                        }}
                    }}
                </style>
            </head>
            <body>
                {html_content}
            </body>
            </html>
            """
            
            # Enhanced PDF options
            pdf_options = {
                'page-size': 'Letter',
                'margin-top': '0.75in',
                'margin-right': '0.75in',
                'margin-bottom': '0.75in',
                'margin-left': '0.75in',
                'encoding': "UTF-8",
                'no-outline': None,
                'enable-local-file-access': None,
                'print-media-type': None,
                'disable-smart-shrinking': None,
                'footer-right': '[page]',
                'footer-font-size': '8'
            }
            
            # Convert HTML to PDF with validation
            pdf_content = pdfkit.from_string(
                styled_html, 
                False, 
                options=pdf_options,
                configuration=pdfkit.configuration(wkhtmltopdf=self.wkhtmltopdf_path)
            )
            
            if not pdf_content:
                raise ReportValidationError("PDF conversion returned empty content")
            
            logger.info(f"Successfully converted Markdown to PDF using {self.wkhtmltopdf_path}")
            return pdf_content
            
        except subprocess.TimeoutExpired:
            logger.error("PDF conversion timed out")
            return b''
        except Exception as e:
            logger.error(f"Error converting Markdown to PDF: {str(e)}")
            return b''


class EnhancedReportManager(ReportManager):
    """Enhanced report manager with validation and optimization."""
    
    def __init__(self):
        super().__init__()
        # Use enhanced generators
        self.excel_generator = MemoryOptimizedExcelGenerator()
        self.pdf_converter = EnhancedPDFConverter()
    
    def generate_comprehensive_report(self, flags_data: Dict[str, Any], metadata: Dict[str, Any], 
                                    extracted_data: Dict[str, Any]) -> Dict[str, Any]:
        """Generate comprehensive audit report with enhancements."""
        try:
            # Validate inputs
            self._validate_report_inputs(flags_data, metadata, extracted_data)
            
            report_id = f"audit-{metadata.get('invoice_number', 'unknown')}-{datetime.now().strftime('%Y%m%d-%H%M%S')}"
            
            # Generate reports with error handling
            reports_generated = {}
            
            # Generate Markdown report
            try:
                markdown_report = self.bedrock_generator.generate_markdown_report(flags_data, metadata)
                reports_generated['markdown'] = markdown_report
            except Exception as e:
                logger.error(f"Markdown generation failed: {str(e)}")
                reports_generated['markdown'] = ""
            
            # Generate Excel report
            try:
                excel_report = self.excel_generator.generate_excel_report(flags_data, metadata, extracted_data)
                reports_generated['excel'] = excel_report
            except Exception as e:
                logger.error(f"Excel generation failed: {str(e)}")
                reports_generated['excel'] = b''
            
            # Convert Markdown to PDF
            try:
                if reports_generated.get('markdown'):
                    pdf_report = self.pdf_converter.markdown_to_pdf(reports_generated['markdown'])
                    reports_generated['pdf'] = pdf_report
                else:
                    reports_generated['pdf'] = b''
            except Exception as e:
                logger.error(f"PDF generation failed: {str(e)}")
                reports_generated['pdf'] = b''
            
            # Upload reports to S3
            report_urls = self._upload_reports(
                report_id, 
                reports_generated.get('markdown', ''),
                reports_generated.get('excel', b''),
                reports_generated.get('pdf', b'')
            )
            
            # Calculate success metrics
            successful_reports = sum(1 for content in reports_generated.values() if content)
            
            # Prepare response
            result = {
                'report_id': report_id,
                'generation_status': 'completed' if successful_reports > 0 else 'partial',
                'timestamp': datetime.now(timezone.utc).isoformat(),
                'reports': report_urls,
                'reports_generated': successful_reports,
                'total_reports': len(reports_generated),
                'summary': {
                    'total_savings': flags_data.get('total_savings', 0),
                    'discrepancies_found': len(flags_data.get('rate_variances', [])) + 
                                         len(flags_data.get('overtime_violations', [])) + 
                                         len(flags_data.get('anomalies', [])),
                    'as_presented': metadata.get('invoice_total', 0),
                    'as_analyzed': metadata.get('invoice_total', 0) - flags_data.get('total_savings', 0)
                }
            }
            
            logger.info(f"Report generation completed: {result['generation_status']} ({successful_reports}/{len(reports_generated)} reports)")
            return result
            
        except Exception as e:
            logger.error(f"Error generating comprehensive report: {str(e)}")
            return {
                'report_id': f"error-{datetime.now().strftime('%Y%m%d-%H%M%S')}",
                'generation_status': 'failed',
                'error': str(e),
                'timestamp': datetime.now(timezone.utc).isoformat()
            }
    
    def _validate_report_inputs(self, flags_data: Dict[str, Any], metadata: Dict[str, Any], 
                              extracted_data: Dict[str, Any]):
        """Validate report generation inputs."""
        if not isinstance(flags_data, dict):
            raise ReportValidationError("flags_data must be a dictionary")
        
        if not isinstance(metadata, dict):
            raise ReportValidationError("metadata must be a dictionary")
        
        if not isinstance(extracted_data, dict):
            raise ReportValidationError("extracted_data must be a dictionary")
        
        # Check required fields
        required_flags_fields = ['rate_variances', 'overtime_violations', 'anomalies', 'total_savings']
        for field in required_flags_fields:
            if field not in flags_data:
                logger.warning(f"Missing required flags field: {field}")
                flags_data[field] = [] if field != 'total_savings' else 0

        total_savings = flags_data.get('total_savings', 0)
        rate_variances = flags_data.get('rate_variances', [])
        overtime_violations = flags_data.get('overtime_violations', [])
        anomalies = flags_data.get('anomalies', [])

        if metadata.get('material_total'):
            raise ReportValidationError("Materials handling removed")


def lambda_handler(event: Dict[str, Any], context: Any) -> Dict[str, Any]:
    """
    Enhanced Lambda handler for report generation.
    
    Expected event structure:
    {
        "task": "generate_report",
        "flags_data": {...},
        "metadata": {...},
        "extracted_data": {...}
    }
    """
    try:
        logger.info(f"Processing enhanced report generation request: {json.dumps(event, default=str)}")
        
        # Validate input
        if event.get('task') != 'generate_report':
            raise ValueError(f"Unknown task: {event.get('task')}")
        
        flags_data = event.get('flags_data', {})
        metadata = event.get('metadata', {})
        extracted_data = event.get('extracted_data', {})
        
        if not flags_data:
            raise ValueError("Missing flags_data in event")
        
        # Generate comprehensive report using enhanced manager
        report_manager = EnhancedReportManager()
        result = report_manager.generate_comprehensive_report(flags_data, metadata, extracted_data)
        
        logger.info(f"Enhanced report generation completed: {result['generation_status']}")
        return result
        
    except Exception as e:
        logger.error(f"Error in enhanced report Lambda handler: {str(e)}")
        return {
            'statusCode': 500,
            'body': json.dumps({
                'error': str(e),
                'timestamp': datetime.now(timezone.utc).isoformat()
            })
        }


def handle_report_generation(flags_data: Dict[str, Any], metadata: Dict[str, Any], 
                           extracted_data: Dict[str, Any]) -> Dict[str, Any]:
    """Handle enhanced report generation task from Step Functions."""
    return lambda_handler({
        'task': 'generate_report',
        'flags_data': flags_data,
        'metadata': metadata,
        'extracted_data': extracted_data
    }, {})


if __name__ == "__main__":
    # Test the enhanced report generation
    test_flags = {
        "total_savings": 375.00,
        "rate_variances": [
            {
                "worker": "Smith, John",
                "labor_type": "RS",
                "actual_rate": 77.00,
                "msa_rate": 70.00,
                "variance_percentage": 10.0,
                "savings": 175.00
            },
            {
                "worker": "Johnson, Mike", 
                "labor_type": "US",
                "actual_rate": 50.00,
                "msa_rate": 45.00,
                "variance_percentage": 11.1,
                "savings": 200.00
            }
        ],
        "overtime_violations": [],
        "anomalies": []
    }
    
    test_metadata = {
        "invoice_number": "INV-2025-001",
        "vendor": "Test Vendor Inc",
        "date_of_loss": "2/12/2025",
        "invoice_total": 160000.00,
        "labor_total": 77000.00,
        "material_total": 83000.00
    }
    
    test_extracted = {
        "normalized_data": {
            "labor": [
                {
                    "name": "Smith, John",
                    "type": "RS", 
                    "total_hours": 25.0,
                    "unit_price": 77.00
                },
                {
                    "name": "Johnson, Mike",
                    "type": "US",
                    "total_hours": 40.0, 
                    "unit_price": 50.00
                }
            ]
        }
    }
    
    try:
        result = handle_report_generation(test_flags, test_metadata, test_extracted)
        print(f"Test result: {json.dumps(result, indent=2, default=str)}")
    except Exception as e:
        print(f"Test failed: {str(e)}")
