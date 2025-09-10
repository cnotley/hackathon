"""
Tests for the Report Generation Lambda function.

This module contains unit tests for report generation functionality including
Bedrock report generation, Excel template processing, PDF conversion, and
comprehensive report workflows.
"""

import json
import pytest
from unittest.mock import Mock, patch, MagicMock
from moto import mock_s3, mock_dynamodb
import boto3
from datetime import datetime, timezone
from decimal import Decimal
from io import BytesIO
import openpyxl

# Import the report Lambda function
import sys
import os
sys.path.append(os.path.join(os.path.dirname(__file__), '..', 'lambda'))

from report_lambda import (
    BedrockReportGenerator,
    ExcelReportGenerator,
    ReportManager
)


class TestBedrockReportGenerator:
    """Test cases for BedrockReportGenerator class."""
    
    def setup_method(self):
        """Set up test fixtures."""
        self.generator = BedrockReportGenerator()
        
        # Sample test data
        self.sample_flags = {
            "total_savings": 375.00,
            "rate_variances": [
                {
                    "worker": "Smith, John",
                    "labor_type": "RS",
                    "actual_rate": 77.00,
                    "msa_rate": 70.00,
                    "variance_percentage": 10.0,
                    "savings": 157.50
                },
                {
                    "worker": "Doe, Jane",
                    "labor_type": "SS",
                    "actual_rate": 99.75,
                    "msa_rate": 95.00,
                    "variance_percentage": 5.0,
                    "savings": 190.00
                }
            ],
            "overtime_violations": [
                {
                    "worker": "Smith, John",
                    "total_hours": 45.0,
                    "overtime_hours": 5.0,
                    "threshold": 40.0
                }
            ],
            "anomalies": [
                {
                    "item": "Safety respirators",
                    "amount": 6313.00,
                    "z_score": 2.5,
                    "description": "Unusually high cost item"
                }
            ]
        }
        
        self.sample_metadata = {
            "invoice_number": "INV-2024-001",
            "vendor": "ABC Construction",
            "date_of_loss": "2/12/2025",
            "invoice_total": 148478.04,
            "labor_total": 77000.00,
            "material_total": 71478.04
        }
    
    @patch('report_lambda.bedrock_client')
    def test_generate_markdown_report_success(self, mock_bedrock_client):
        """Test successful Markdown report generation using Bedrock."""
        # Mock Bedrock response
        mock_response = {
            'body': Mock()
        }
        mock_response['body'].read.return_value = json.dumps({
            'content': [{
                'text': '# MSA Audit Report\n\n## Executive Summary\n\nThis audit found $375.00 in potential savings.'
            }]
        }).encode()
        
        mock_bedrock_client.invoke_model.return_value = mock_response
        
        # Generate report
        result = self.generator.generate_markdown_report(self.sample_flags, self.sample_metadata)
        
        # Assertions
        assert result is not None
        assert '# MSA Audit Report' in result
        assert '$375.00' in result
        mock_bedrock_client.invoke_model.assert_called_once()
    
    @patch('report_lambda.bedrock_client')
    def test_generate_markdown_report_fallback(self, mock_bedrock_client):
        """Test fallback report generation when Bedrock fails."""
        # Mock Bedrock to raise exception
        mock_bedrock_client.invoke_model.side_effect = Exception("Bedrock error")
        
        # Generate report (should use fallback)
        result = self.generator.generate_markdown_report(self.sample_flags, self.sample_metadata)
        
        # Assertions
        assert result is not None
        assert '# MSA Audit Report' in result
        assert 'Executive Summary' in result
        assert '$375.00' in result
        assert 'Smith, John' in result
        assert 'Rate Variances (2 found)' in result
        assert 'Overtime Violations (1 found)' in result
    
    def test_build_report_prompt(self):
        """Test building the Bedrock prompt."""
        prompt = self.generator._build_report_prompt(self.sample_flags, self.sample_metadata)
        
        # Verify prompt contains key information
        assert 'INV-2024-001' in prompt
        assert 'ABC Construction' in prompt
        assert '$148,478.04' in prompt
        assert '$375.00' in prompt
        assert 'Smith, John' in prompt
        assert 'RS' in prompt
        assert '$77.00' in prompt
        assert '$70.00' in prompt
    
    def test_generate_fallback_report_structure(self):
        """Test the structure of the fallback report."""
        result = self.generator._generate_fallback_report(self.sample_flags, self.sample_metadata)
        
        # Check report structure
        assert '# MSA Audit Report' in result
        assert '## Executive Summary' in result
        assert '## Financial Summary' in result
        assert '## Detailed Findings' in result
        assert '### Rate Variances' in result
        assert '### Overtime Violations' in result
        assert '### Cost Anomalies' in result
        assert '## Recommendations' in result
        assert '## Compliance Assessment' in result
        
        # Check specific content
        assert 'Total Discrepancies: 4' in result  # 2 rate + 1 overtime + 1 anomaly
        assert 'Potential Savings: $375.00' in result
        assert 'Safety respirators' in result


class TestExcelReportGenerator:
    """Test cases for ExcelReportGenerator class."""
    
    def setup_method(self):
        """Set up test fixtures."""
        self.generator = ExcelReportGenerator()
        
        # Sample test data
        self.sample_flags = {
            "total_savings": 375.00,
            "rate_variances": [
                {
                    "worker": "Smith, John",
                    "labor_type": "RS",
                    "actual_rate": 77.00,
                    "msa_rate": 70.00,
                    "variance_percentage": 10.0,
                    "savings": 157.50
                }
            ],
            "overtime_violations": [],
            "anomalies": []
        }
        
        self.sample_metadata = {
            "invoice_number": "INV-2024-001",
            "vendor": "ABC Construction",
            "date_of_loss": "2/12/2025",
            "invoice_total": 148478.04,
            "labor_total": 77000.00,
            "material_total": 71478.04,
            "page_count": 22,
            "file_size": 1024000
        }
        
        self.sample_extracted_data = {
            "normalized_data": {
                "labor": [
                    {
                        "name": "Smith, John",
                        "type": "RS",
                        "unit_price": 77.00,
                        "total_hours": 35.0,
                        "total_cost": 2695.00
                    }
                ]
            }
        }
    
    @mock_s3
    def test_generate_excel_report_fallback(self):
        """Test Excel report generation with fallback (no template)."""
        # Mock S3 to return empty template (triggers fallback)
        with patch.object(self.generator, '_download_template', return_value=b''):
            result = self.generator.generate_excel_report(
                self.sample_flags, 
                self.sample_metadata, 
                self.sample_extracted_data
            )
        
        # Verify Excel content was generated
        assert result is not None
        assert len(result) > 0
        
        # Load and verify Excel content
        workbook = openpyxl.load_workbook(BytesIO(result))
        
        # Check that sheets were created
        assert 'Project Information' in workbook.sheetnames
        assert 'Project Summary' in workbook.sheetnames
        assert 'Labor Export' in workbook.sheetnames
    
    def test_fill_project_information(self):
        """Test filling Project Information sheet."""
        workbook = openpyxl.Workbook()
        
        self.generator._fill_project_information(workbook, self.sample_metadata)
        
        # Verify sheet was created and populated
        assert 'Project Information' in workbook.sheetnames
        ws = workbook['Project Information']
        
        # Check headers
        assert ws.cell(row=1, column=1).value == 'Field'
        assert ws.cell(row=1, column=2).value == 'Value'
        
        # Check data
        invoice_number_found = False
        vendor_found = False
        for row in range(2, ws.max_row + 1):
            field = ws.cell(row=row, column=1).value
            value = ws.cell(row=row, column=2).value
            
            if field == 'Invoice Number':
                assert value == 'INV-2024-001'
                invoice_number_found = True
            elif field == 'Vendor':
                assert value == 'ABC Construction'
                vendor_found = True
        
        assert invoice_number_found
        assert vendor_found
    
    def test_fill_project_summary(self):
        """Test filling Project Summary sheet."""
        workbook = openpyxl.Workbook()
        
        self.generator._fill_project_summary(workbook, self.sample_flags, self.sample_metadata)
        
        # Verify sheet was created and populated
        assert 'Project Summary' in workbook.sheetnames
        ws = workbook['Project Summary']
        
        # Check headers
        assert ws.cell(row=1, column=1).value == 'Category'
        assert ws.cell(row=1, column=2).value == 'As Presented'
        assert ws.cell(row=1, column=3).value == 'As Analyzed'
        assert ws.cell(row=1, column=4).value == 'Savings'
        
        # Check labor costs row
        assert ws.cell(row=2, column=1).value == 'Labor Costs'
        assert '$77,000.00' in str(ws.cell(row=2, column=2).value)
        assert '$76,625.00' in str(ws.cell(row=2, column=3).value)  # 77000 - 375
        assert '$375.00' in str(ws.cell(row=2, column=4).value)
        
        # Check total project row
        total_row = ws.max_row
        assert ws.cell(row=total_row, column=1).value == 'Total Project'
        assert '$148,478.04' in str(ws.cell(row=total_row, column=2).value)
    
    def test_fill_labor_export(self):
        """Test filling Labor Export sheet."""
        workbook = openpyxl.Workbook()
        
        self.generator._fill_labor_export(workbook, self.sample_extracted_data, self.sample_flags)
        
        # Verify sheet was created and populated
        assert 'Labor Export' in workbook.sheetnames
        ws = workbook['Labor Export']
        
        # Check headers
        expected_headers = ['Worker Name', 'Labor Type', 'Hours', 'Rate', 'Total', 'MSA Rate', 'Variance', 'Savings']
        for col_idx, expected_header in enumerate(expected_headers, 1):
            assert ws.cell(row=1, column=col_idx).value == expected_header
        
        # Check data row
        assert ws.cell(row=2, column=1).value == 'Smith, John'
        assert ws.cell(row=2, column=2).value == 'RS'
        assert '35.0' in str(ws.cell(row=2, column=3).value)
        assert '$77.00' in str(ws.cell(row=2, column=4).value)


class TestReportManager:
    """Test cases for ReportManager class."""
    
    def setup_method(self):
        """Set up test fixtures."""
        self.manager = ReportManager()
        
        # Sample comprehensive test data
        self.sample_flags = {
            "total_savings": 375.00,
            "rate_variances": [
                {
                    "worker": "Smith, John",
                    "labor_type": "RS",
                    "actual_rate": 77.00,
                    "msa_rate": 70.00,
                    "variance_percentage": 10.0,
                    "savings": 157.50
                }
            ],
            "overtime_violations": [
                {
                    "worker": "Smith, John",
                    "total_hours": 45.0,
                    "overtime_hours": 5.0,
                    "threshold": 40.0
                }
            ],
            "anomalies": []
        }
        
        self.sample_metadata = {
            "invoice_number": "INV-2024-001",
            "vendor": "ABC Construction",
            "date_of_loss": "2/12/2025",
            "invoice_total": 148478.04,
            "labor_total": 77000.00,
            "material_total": 71478.04
        }
        
        self.sample_extracted_data = {
            "normalized_data": {
                "labor": [
                    {
                        "name": "Smith, John",
                        "type": "RS",
                        "unit_price": 77.00,
                        "total_hours": 45.0,
                        "total_cost": 3465.00
                    }
                ]
            }
        }
    
    @mock_s3
    @patch('report_lambda.BedrockReportGenerator')
    @patch('report_lambda.ExcelReportGenerator')
    def test_generate_comprehensive_report_success(self, mock_excel_generator_cls, mock_bedrock_generator_cls):
        """Test successful comprehensive report generation."""
        # Setup S3 mock
        s3_client = boto3.client('s3', region_name='us-east-1')
        s3_client.create_bucket(Bucket='msa-audit-reports')
        
        # Mock generators
        mock_bedrock_instance = Mock()
        mock_excel_instance = Mock()
        mock_bedrock_generator_cls.return_value = mock_bedrock_instance
        mock_excel_generator_cls.return_value = mock_excel_instance

        # Configure mock responses
        mock_bedrock_instance.generate_markdown_report.return_value = "# Test Report\n\nContent here"
        mock_excel_instance.generate_excel_report.return_value = b'Excel content'

        # Mock S3 uploads
        with patch.object(self.manager, '_upload_reports') as mock_upload:
            mock_upload.return_value = {
                'markdown': 's3://bucket/report.md',
                'excel': 's3://bucket/report.xlsx'
            }
            
            result = self.manager.generate_comprehensive_report(
                self.sample_flags, 
                self.sample_metadata, 
                self.sample_extracted_data
            )
        
        # Assertions
        assert result['generation_status'] == 'completed'
        assert 'report_id' in result
        assert result['summary']['total_savings'] == 375.00
        assert result['summary']['discrepancies_found'] == 2  # 1 rate variance + 1 overtime
        assert result['summary']['as_presented'] == 148478.04
        assert result['summary']['as_analyzed'] == 148103.04  # 148478.04 - 375.00
        
        # Verify generators were called
        mock_bedrock_instance.generate_markdown_report.assert_called_once()
        mock_excel_instance.generate_excel_report.assert_called_once()
    
    @patch('report_lambda.BedrockReportGenerator')
    def test_generate_comprehensive_report_error(self, mock_bedrock_generator):
        """Test comprehensive report generation error handling."""
        # Mock generator to raise exception
        mock_bedrock_instance = Mock()
        mock_bedrock_generator.return_value = mock_bedrock_instance
        mock_bedrock_instance.generate_markdown_report.side_effect = Exception("Generation failed")
        
        result = self.manager.generate_comprehensive_report(
            self.sample_flags, 
            self.sample_metadata, 
            self.sample_extracted_data
        )
        
        # Should return error status
        assert result['generation_status'] == 'failed'
        assert 'error' in result
        assert 'Generation failed' in result['error']
    
    @mock_s3
    def test_upload_reports(self):
        """Test report upload to S3."""
        # Setup S3 mock
        s3_client = boto3.client('s3', region_name='us-east-1')
        s3_client.create_bucket(Bucket='msa-audit-reports')
        
        # Test data
        report_id = "test-report-123"
        markdown_content = "# Test Report"
        excel_content = b'Excel content'
        
        # Upload reports
        with patch.dict(os.environ, {'REPORTS_BUCKET': 'msa-audit-reports'}):
            result = self.manager._upload_reports(report_id, markdown_content, excel_content, None)
        
        # Verify uploads
        assert 'markdown' in result
        assert 'excel' in result
        assert 'pdf' not in result # PDF is no longer generated
        
        # Verify S3 objects were created
        objects = s3_client.list_objects_v2(Bucket='msa-audit-reports')
        assert objects['KeyCount'] == 2 # Markdown, Excel
        
        keys = [obj['Key'] for obj in objects['Contents']]
        assert f'reports/{report_id}/{report_id}.md' in keys
        assert f'reports/{report_id}/{report_id}.xlsx' in keys


class TestLambdaHandler:
    """Test cases for the main Lambda handler."""
    
    def setup_method(self):
        """Set up test fixtures."""
        self.sample_event = {
            "task": "generate_report",
            "flags_data": {
                "total_savings": 375.00,
                "rate_variances": [
                    {
                        "worker": "Smith, John",
                        "labor_type": "RS",
                        "actual_rate": 77.00,
                        "msa_rate": 70.00,
                        "variance_percentage": 10.0,
                        "savings": 157.50
                    }
                ],
                "overtime_violations": [],
                "anomalies": []
            },
            "metadata": {
                "invoice_number": "INV-2024-001",
                "vendor": "ABC Construction",
                "invoice_total": 148478.04
            },
            "extracted_data": {
                "normalized_data": {
                    "labor": [
                        {
                            "name": "Smith, John",
                            "type": "RS",
                            "unit_price": 77.00,
                            "total_hours": 35.0
                        }
                    ]
                }
            }
        }
    
    @patch('report_lambda.ReportManager')
    def test_lambda_handler_success(self, mock_report_manager):
        """Test successful Lambda handler execution."""
        # Mock ReportManager
        mock_manager_instance = Mock()
        mock_report_manager.return_value = mock_manager_instance
        
        mock_manager_instance.generate_comprehensive_report.return_value = {
            'report_id': 'test-report-123',
            'generation_status': 'completed',
            'timestamp': '2025-01-09T20:50:00Z',
            'summary': {
                'total_savings': 375.00,
                'discrepancies_found': 1,
                'as_presented': 148478.04,
                'as_analyzed': 148103.04
            }
        }
        
        # Execute handler
        result = lambda_handler(self.sample_event, {})
        
        # Assertions
        assert result['generation_status'] == 'completed'
        assert result['report_id'] == 'test-report-123'
        assert result['summary']['total_savings'] == 375.00
        
        # Verify manager was called correctly
        mock_manager_instance.generate_comprehensive_report.assert_called_once()
        call_args = mock_manager_instance.generate_comprehensive_report.call_args[0]
        assert call_args[0] == self.sample_event['flags_data']
        assert call_args[1] == self.sample_event['metadata']
        assert call_args[2] == self.sample_event['extracted_data']
    
    def test_lambda_handler_unknown_task(self):
        """Test Lambda handler with unknown task."""
        event = {
            "task": "unknown_task",
            "flags_data": {}
        }
        
        result = lambda_handler(event, {})
        
        # Should return error
        assert result['statusCode'] == 500
        body = json.loads(result['body'])
        assert 'Unknown task' in body['error']
    
    def test_lambda_handler_missing_flags_data(self):
        """Test Lambda handler with missing flags_data."""
        event = {
            "task": "generate_report",
            "metadata": {},
            "extracted_data": {}
        }
        
        result = lambda_handler(event, {})
        
        # Should return error
        assert result['statusCode'] == 500
        body = json.loads(result['body'])
        assert 'Missing flags_data' in body['error']
    
    @patch('report_lambda.ReportManager')
    def test_handle_report_generation(self, mock_report_manager):
        """Test handle_report_generation function."""
        # Mock ReportManager
        mock_manager_instance = Mock()
        mock_report_manager.return_value = mock_manager_instance
        
        mock_manager_instance.generate_comprehensive_report.return_value = {
            'report_id': 'test-report-456',
            'generation_status': 'completed'
        }
        
        # Test data
        flags_data = {"total_savings": 100.00}
        metadata = {"invoice_number": "TEST-001"}
        extracted_data = {"normalized_data": {}}
        
        # Execute function
        result = handle_report_generation(flags_data, metadata, extracted_data)
        
        # Assertions
        assert result['generation_status'] == 'completed'
        assert result['report_id'] == 'test-report-456'


class TestIntegrationScenarios:
    """Integration test scenarios for complete report generation workflows."""
    
    @mock_s3
    @patch('report_lambda.bedrock_client')
    def test_complete_report_generation_workflow(self, mock_bedrock_client):
        """Test complete report generation workflow with all components."""
        # Setup S3 mock
        s3_client = boto3.client('s3', region_name='us-east-1')
        s3_client.create_bucket(Bucket='msa-audit-reports')
        
        # Mock Bedrock response
        mock_bedrock_response = {
            'body': Mock()
        }
        mock_bedrock_response['body'].read.return_value = json.dumps({
            'content': [{
                'text': '# MSA Audit Report\n\n## Executive Summary\n\nOvercharge: $375 on RS labor; Savings: 10% of $160k'
            }]
        }).encode()
        mock_bedrock_client.invoke_model.return_value = mock_bedrock_response
        
        # Mock PDF generation
        
        # Comprehensive test data matching requirements
        flags_data = {
            "total_savings": 375.00,
            "rate_variances": [
                {
                    "worker": "Smith, John",
                    "labor_type": "RS",
                    "actual_rate": 77.00,
                    "msa_rate": 70.00,
                    "variance_percentage": 10.0,
                    "savings": 375.00,
                    "flag": "Rate variance exceeds 5% threshold"
                }
            ],
            "overtime_violations": [],
            "anomalies": []
        }
        
        metadata = {
            "invoice_number": "INV-2024-001",
            "vendor": "ABC Construction",
            "date_of_loss": "2/12/2025",
            "invoice_total": 148478.04,
            "labor_total": 77000.00,
            "material_total": 71478.04,
            "page_count": 22,
            "file_size": 1024000
        }
        
        extracted_data = {
            "normalized_data": {
                "labor": [
                    {
                        "name": "Smith, John",
                        "type": "RS",
                        "unit_price": 77.00,
                        "total_hours": 35.0,
                        "total_cost": 2695.00
                    }
                ]
            }
        }
        
        # Execute complete workflow
        with patch.dict(os.environ, {'REPORTS_BUCKET': 'msa-audit-reports'}):
            manager = ReportManager()
            result = manager.generate_comprehensive_report(flags_data, metadata, extracted_data)
        
        # Verify comprehensive results
        assert result['generation_status'] == 'completed'
        assert 'report_id' in result
        assert result['summary']['total_savings'] == 375.00
        assert result['summary']['as_presented'] == 148478.04
        assert result['summary']['as_analyzed'] == 148103.04  # 148478.04 - 375.00
        
        # Verify Bedrock was called for report generation
        mock_bedrock_client.invoke_model.assert_called_once()
        
        # Verify PDF generation was attempted
        
        # Verify S3 uploads occurred
        objects = s3_client.list_objects_v2(Bucket='msa-audit-reports')
        assert objects['KeyCount'] == 2  # Markdown, Excel
    
    def test_excel_summary_total_matches_requirements(self):
        """Test that Excel summary total matches requirements ($148,478.04 adjusted)."""
        generator = ExcelReportGenerator()
        
        # Test data matching requirements
        flags_data = {"total_savings": 375.00}
        metadata = {
            "labor_total": 77000.00,
            "material_total": 71478.04,
            "invoice_total": 148478.04
        }
        extracted_data = {"normalized_data": {"labor": []}}
        
        # Generate Excel report
        excel_content = generator._generate_fallback_excel(flags_data, metadata, extracted_data)
        
        # Load and verify Excel content
        workbook = openpyxl.load_workbook(BytesIO(excel_content))
        ws = workbook['Project Summary']
        
        # Find total project row
        total_row = None
        for row in range(1, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == 'Total Project':
                total_row = row
                break
        
        assert total_row is not None
        
        # Verify totals match requirements
        as_presented = ws.cell(row=total_row, column=2).value
        as_analyzed = ws.cell(row=total_row, column=3).value
        savings = ws.cell(row=total_row, column=4).value
        
        # Extract numeric values from formatted strings
        assert '$148,478.04' in str(as_presented)
        assert '$148,103.04' in str(as_analyzed)  # 148478.04 - 375.00
        assert '$375.00' in str(savings)

    def test_project_summary_totals_match_requirements(self):
        generator = ExcelReportGenerator()
        flags = {"total_savings": 375.0}
        metadata = {
            "invoice_total": 148478.04,
            "labor_total": 77000.0,
            "material_total": 71478.04,
        }
        extracted = {"normalized_data": {"labor": []}}

        workbook_bytes = generator._generate_fallback_excel(flags, metadata, extracted)
        workbook = openpyxl.load_workbook(BytesIO(workbook_bytes))
        summary_sheet = workbook['Project Summary']

        totals = {}
        for row in range(2, summary_sheet.max_row + 1):
            category = summary_sheet.cell(row=row, column=1).value
            if category == 'Total Project':
                totals['as_presented'] = summary_sheet.cell(row=row, column=2).value
                totals['as_analyzed'] = summary_sheet.cell(row=row, column=3).value
                totals['savings'] = summary_sheet.cell(row=row, column=4).value
                break

        assert totals
        assert '$148,478.04' in str(totals['as_presented'])
        assert '$148,103.04' in str(totals['as_analyzed'])
        assert '$375.00' in str(totals['savings'])


class TestErrorHandling:
    """Test cases for error handling scenarios."""
    
    def test_bedrock_generator_with_invalid_data(self):
        """Test Bedrock generator with invalid input data."""
        generator = BedrockReportGenerator()
        
        # Test with empty/invalid data
        empty_flags = {}
        empty_metadata = {}
        
        # Should not raise exception, should use fallback
        result = generator._generate_fallback_report(empty_flags, empty_metadata)
        
        assert result is not None
        assert '# MSA Audit Report' in result
        assert 'Total Discrepancies: 0' in result
    
    def test_excel_generator_with_missing_data(self):
        """Test Excel generator with missing data fields."""
        generator = ExcelReportGenerator()
        
        # Test with minimal data
        minimal_flags = {"total_savings": 0}
        minimal_metadata = {}
        minimal_extracted = {"normalized_data": {"labor": []}}
        
        # Should not raise exception
        result = generator._generate_fallback_excel(
            minimal_flags,
            minimal_metadata,
            minimal_extracted
        )
        assert result is not None
        workbook = openpyxl.load_workbook(BytesIO(result))
        assert 'Project Summary' in workbook.sheetnames

    def test_excel_report_contains_as_analyzed_column(self):
        generator = ExcelReportGenerator()
        flags = {"total_savings": 375.0}
        metadata = {"invoice_total": 148478.04, "labor_total": 77000.0}
        extracted = {"normalized_data": {"labor": []}}

        workbook_bytes = generator._generate_fallback_excel(flags, metadata, extracted)
        workbook = openpyxl.load_workbook(BytesIO(workbook_bytes))
        summary_sheet = workbook['Project Summary']

        headers = [summary_sheet.cell(row=1, column=col).value for col in range(1, 5)]
        assert headers[2] == 'As Analyzed'
        labor_row = summary_sheet.cell(row=2, column=3).value
        assert '$76,625.00' in str(labor_row)
