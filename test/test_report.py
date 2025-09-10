import importlib
generate_report = importlib.import_module("lambda.report_lambda").generate_report

def test_report_outputs():
    extr = {"invoice_number":"3034894","project":"Liverpool","loss_date":"02/12/2025","summary":{"labor":77000},"labor":[{"name":"Manderville","type":"Restoration Specialist","code":"RS","rate":77,"total_hours":55,"total":4812}]}
    comp = {"flags":[{"type":"rate_high_vs_mwo","code":"RS","expected":70,"seen":77}], "estimated_savings": 16000}
    out = generate_report(extr, comp, out_bucket=None)
    assert "report.xlsx" in out["generated"]
    assert "report.md" in out["generated"]
    import openpyxl, io
    wb = openpyxl.load_workbook(io.BytesIO(out["files"]["report.xlsx"]))
    ws = wb["Project Summary"]
    holds = [ws.cell(r,4).value for r in range(2, ws.max_row+1)]
    assert any(h > 0 for h in holds)
